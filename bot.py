from telegram import Update, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# 🔐 BOT TOKENINGIZNI SHU YERGA QO'YING
TOKEN = "8341816131:AAGKLBRlUvw-VF4YJ9wB9Fi89ehNd-LsImM"

# 📁 Fayllar
FAYL_FOYDALANUVCHI = "data.xlsx"
FAYL_DAVOMAT = "saqla.xlsx"
FAYL_MASTERLAR = "masterlar.xlsx"

# 👥 Ishchilar ro‘yxati
ISHCHILAR = ["Ibrohimjon", "Abror", "Zarifa", "Gulnoza", "Yulduzoy", "Zulfizar"]
MASTERLAR = ["Biloldin", "Xolidaxola", "NafisaMaster", "Nafisa", "Madina", "Gulshanoy", "Dildora"]

# 🔑 Login bosqichlari
LOGIN_USERNAME, LOGIN_PASSWORD = range(2)

# 📘 Fayllarni tayyorlash
def tayyorla_fayllar():
    if not os.path.exists(FAYL_FOYDALANUVCHI):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Password"])
        ws.append(["admin", "7777"])
        ws.append(["bichuv", "333"])
        ws.append(["masterlar", "555"])
        ws.append(["ibroxim", "999"])
        wb.save(FAYL_FOYDALANUVCHI)

    if not os.path.exists(FAYL_DAVOMAT):
        wb = Workbook()
        ws = wb.active
        ws.title = "Davomat"
        ws.append(["Sana", "Ishchi", "Holat", "Username"])
        wb.save(FAYL_DAVOMAT)

    if not os.path.exists(FAYL_MASTERLAR):
        wb = Workbook()
        ws = wb.active
        ws.title = "Masterlar"
        ws.append(["Sana", "Ishchi", "Holat", "Username"])
        wb.save(FAYL_MASTERLAR)

# 🔐 Login tekshirish
def tekshir_login(username, password):
    wb = load_workbook(FAYL_FOYDALANUVCHI)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

# 🧾 Bugun yozilganmi
def bugun_yozilganmi(ism, fayl=FAYL_DAVOMAT):
    if not os.path.exists(fayl):
        return False
    wb = load_workbook(fayl)
    ws = wb.active
    bugun = datetime.now().strftime("%Y-%m-%d")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == ism and str(row[0]).startswith(bugun):
            return True
    return False

# 📝 Davomat yozish
def yoz_davomat(ism, holat, username, fayl=FAYL_DAVOMAT):
    bugun = datetime.now()
    if bugun.weekday() == 6:  # Yakshanba
        return "❌ Bugun yakshanba, davomat olinmaydi."
    wb = load_workbook(fayl)
    ws = wb.active
    sana = bugun.strftime("%Y-%m-%d %H:%M:%S")
    ws.append([sana, ism, holat, username])
    wb.save(fayl)
    return f"✅ Saqlandi: {ism} - {holat}"

# 🧾 Hozirgi oy bo‘yicha davomat hisobot
def davomat_hisobot(fayl=FAYL_DAVOMAT, ishchilar=ISHCHILAR):
    if not os.path.exists(fayl):
        return "❌ Ma’lumotlar topilmadi."
    wb = load_workbook(fayl)
    ws = wb.active
    oy = datetime.now().strftime("%m")
    natija = f"📊 Hozirgi oy davomat hisobot:\n\n"
    kunlar = {ism: 0 for ism in ishchilar}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sana, ism, holat, _ = row
        if sana[5:7] == oy:
            if ism in kunlar and holat == "Keldi":
                kunlar[ism] += 1
    for ism, son in kunlar.items():
        natija += f"{ism}: {son} kun keldi\n"
    return natija

# 🔰 /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [KeyboardButton("👕 Bichuv bo‘limi"), KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
        [KeyboardButton("📥 Ma'lumotlarni yuklash")],
        [KeyboardButton("/login"), KeyboardButton("/logout")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Assalomu alaykum! /login orqali tizimga kiring:", reply_markup=reply_markup)

# 🔐 Login bosqichlari
async def login_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Login kiriting:")
    return LOGIN_USERNAME

async def login_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tmp_username'] = update.message.text.strip()
    await update.message.reply_text("Parolni kiriting:")
    return LOGIN_PASSWORD

async def login_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = context.user_data.get('tmp_username')
    password = update.message.text.strip()
    if tekshir_login(username, password):
        context.user_data['auth'] = True
        context.user_data['username'] = username

        # 🟢 Menyularni foydalanuvchiga moslash
        if username == "admin":
            keyboard = [
                [KeyboardButton("👕 Bichuv bo‘limi"), KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                [KeyboardButton("📥 Ma'lumotlarni yuklash")],
                [KeyboardButton("📊 Davomat hisobot")],
                [KeyboardButton("/logout")]
            ]
        elif username == "bichuv":
            keyboard = [
                [KeyboardButton("👕 Bichuv bo‘limi")],
                [KeyboardButton("/logout")]
            ]
        elif username == "masterlar":
            keyboard = [
                [KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                [KeyboardButton("/logout")]
            ]
        else:
            keyboard = [
                [KeyboardButton("/logout")]
            ]

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(f"✅ Tizimga kirdingiz: {username}", reply_markup=reply_markup)
    else:
        await update.message.reply_text("❌ Login yoki parol noto‘g‘ri.")
    return ConversationHandler.END

# 🚪 Logout
async def logout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("auth"):
        foydalanuvchi = context.user_data.get("username", "")
        context.user_data.clear()
        await update.message.reply_text(f"🚪 {foydalanuvchi} tizimdan chiqdi.")
    else:
        await update.message.reply_text("Siz tizimda emassiz.")

# 📩 Asosiy xabarlar
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    auth = context.user_data.get("auth", False)
    username = context.user_data.get("username", "")

    # 👕 Bichuv bo‘limi
    if text == "👕 Bichuv bo‘limi":
        if not auth:
            await update.message.reply_text("Iltimos, avval /login orqali tizimga kiring.")
            return
        if datetime.now().weekday() == 6:
            await update.message.reply_text("🛑 Bugun yakshanba. Davomat olinmaydi.")
            return
        buttons = [[KeyboardButton(ism)] for ism in ISHCHILAR]
        buttons.append([KeyboardButton("🔙 Orqaga")])
        reply_markup = ReplyKeyboardMarkup(buttons, resize_keyboard=True)
        await update.message.reply_text("Ishchini tanlang:", reply_markup=reply_markup)

    # 👨‍🏫 Masterlar bo‘limi
    elif text == "👨‍🏫 Masterlar bo‘limi":
        if not auth:
            await update.message.reply_text("Iltimos, avval /login orqali tizimga kiring.")
            return
        if datetime.now().weekday() == 6:
            await update.message.reply_text("🛑 Bugun yakshanba. Davomat olinmaydi.")
            return
        buttons = [[KeyboardButton(ism)] for ism in MASTERLAR]
        buttons.append([KeyboardButton("🔙 Orqaga")])
        reply_markup = ReplyKeyboardMarkup(buttons, resize_keyboard=True)
        await update.message.reply_text("Masterni tanlang:", reply_markup=reply_markup)

    # 📥 Ma'lumotlarni yuklash (faqat admin)
    elif text == "📥 Ma'lumotlarni yuklash":
        if not auth or username != "admin":
            await update.message.reply_text("🚫 Bu bo‘lim faqat admin uchun.")
            return
        oylar = [
            ["Bichuv ma’lumotlari", "Masterlar ma’lumotlari"],
            ["🔙 Orqaga"]
        ]
        reply_markup = ReplyKeyboardMarkup(oylar, resize_keyboard=True)
        await update.message.reply_text("Qaysi bo‘lim ma’lumotlarini ko‘rmoqchisiz?", reply_markup=reply_markup)

    elif text in ["Bichuv ma’lumotlari", "Masterlar ma’lumotlari"]:
        if not auth or username != "admin":
            await update.message.reply_text("🚫 Bu bo‘lim faqat admin uchun.")
            return
        fayl = FAYL_DAVOMAT if text == "Bichuv ma’lumotlari" else FAYL_MASTERLAR
        if not os.path.exists(fayl):
            await update.message.reply_text("🚫 Fayl topilmadi.")
            return
        wb = load_workbook(fayl)
        ws = wb.active
        natija = f"📄 {text}:\n\n"
        for row in ws.iter_rows(min_row=2, values_only=True):
            sana, ism, holat, foydalanuvchi = row
            natija += f"{sana} — {ism} — {holat} — {foydalanuvchi}\n"
        await update.message.reply_text(natija)

    # 📊 Davomat hisobot
    elif text == "📊 Davomat hisobot":
        if not auth or username != "admin":
            await update.message.reply_text("🚫 Bu bo‘lim faqat admin uchun.")
            return
        keyboard = [
            [KeyboardButton("Bichuv davomati")],
            [KeyboardButton("Masterlar davomati")],
            [KeyboardButton("🔙 Orqaga")]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Qaysi bo‘lim davomatini ko‘rmoqchisiz?", reply_markup=reply_markup)

    elif text == "Bichuv davomati":
        javob = davomat_hisobot(FAYL_DAVOMAT, ISHCHILAR)
        await update.message.reply_text(javob)

    elif text == "Masterlar davomati":
        javob = davomat_hisobot(FAYL_MASTERLAR, MASTERLAR)
        await update.message.reply_text(javob)

    # 🔙 Orqaga
    elif text == "🔙 Orqaga":
        if auth and username == "admin":
            keyboard = [
                [KeyboardButton("👕 Bichuv bo‘limi"), KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                [KeyboardButton("📥 Ma'lumotlarni yuklash")],
                [KeyboardButton("📊 Davomat hisobot")],
                [KeyboardButton("/logout")]
            ]
        elif auth and username == "bichuv":
            keyboard = [
                [KeyboardButton("👕 Bichuv bo‘limi")],
                [KeyboardButton("/logout")]
            ]
        elif auth and username == "masterlar":
            keyboard = [
                [KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                [KeyboardButton("/logout")]
            ]
        else:
            keyboard = [
                [KeyboardButton("👕 Bichuv bo‘limi"), KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                [KeyboardButton("/login"), KeyboardButton("/logout")]
            ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Asosiy menyuga qaytdingiz.", reply_markup=reply_markup)

    # Ishchi tanlansa
    elif text in ISHCHILAR + MASTERLAR:
        if not auth:
            await update.message.reply_text("❗ Avval /login orqali tizimga kiring.")
            return
        fayl = FAYL_MASTERLAR if text in MASTERLAR else FAYL_DAVOMAT
        if bugun_yozilganmi(text, fayl):
            await update.message.reply_text(f"⚠️ {text} bugun uchun kiritilgan.")
            return
        buttons = [
            [KeyboardButton(f"✅ {text} keldi"), KeyboardButton(f"❌ {text} kelmadi")],
            [KeyboardButton("🔙 Orqaga")]
        ]
        reply_markup = ReplyKeyboardMarkup(buttons, resize_keyboard=True)
        await update.message.reply_text(f"{text} uchun holatni tanlang:", reply_markup=reply_markup)

    # ✅ yoki ❌ bosilganda
    elif text.startswith("✅") or text.startswith("❌"):
        if not auth:
            await update.message.reply_text("❗ Avval /login orqali tizimga kiring.")
            return
        parts = text.split(" ", 2)
        if len(parts) >= 3:
            holat = "Keldi" if text.startswith("✅") else "Kelmadi"
            ism = parts[1]
            fayl = FAYL_MASTERLAR if ism in MASTERLAR else FAYL_DAVOMAT
            javob = yoz_davomat(ism, holat, username, fayl)
            if auth and username == "admin":
                keyboard = [
                    [KeyboardButton("👕 Bichuv bo‘limi"), KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                    [KeyboardButton("📥 Ma'lumotlarni yuklash")],
                    [KeyboardButton("📊 Davomat hisobot")],
                    [KeyboardButton("/logout")]
                ]
            elif auth and username == "bichuv":
                keyboard = [
                    [KeyboardButton("👕 Bichuv bo‘limi")],
                    [KeyboardButton("/logout")]
                ]
            elif auth and username == "masterlar":
                keyboard = [
                    [KeyboardButton("👨‍🏫 Masterlar bo‘limi")],
                    [KeyboardButton("/logout")]
                ]
            else:
                keyboard = [
                    [KeyboardButton("/logout")]
                ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.message.reply_text(javob, reply_markup=reply_markup)

    else:
        await update.message.reply_text("Iltimos, menyudan tanlang yoki /login orqali tizimga kiring.")

# 🔧 Botni ishga tushirish
def main():
    tayyorla_fayllar()
    app = ApplicationBuilder().token(TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("login", login_start)],
        states={
            LOGIN_USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, login_username)],
            LOGIN_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, login_password)],
        },
        fallbacks=[],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("logout", logout))
    app.add_handler(conv_handler)
    app.add_handler(MessageHandler(filters.TEXT, message_handler))

    print("🤖 Bot ishga tushdi...")
    app.run_polling()

if __name__ == "__main__":
    main()
